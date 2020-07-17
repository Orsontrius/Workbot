import openpyxl
import re

file = "16-22 July 2020_v1.xlsx"

def parse_spreadsheet(file):
    wb = openpyxl.load_workbook(filename=file)
    ws = wb.active

    #get the row numbers for the date and my row
    dateRow = 3
    cjRow = 0
    x = 1
    for row in ws["A"]:
        if row.value == "CJ":
            cjRow = x 
        x += 1

    for i in ws.iter_rows(min_row=cjRow, max_row=cjRow, values_only=True):
        times = i
    for i in ws.iter_rows(min_row=dateRow, max_row=dateRow, values_only=True):
        dates = i


    for i in range(len(times)):
        if re.search("[am]|[pm]", str(times[i])):
            

    


    

parse_spreadsheet(file)